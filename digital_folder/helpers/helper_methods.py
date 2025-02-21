from typing import Type, Any, List, Optional
from uuid import UUID

import pandas as pd
from openpyxl.reader.excel import load_workbook
from pydantic import SecretStr, BaseModel, create_model


def get_scrt_key(db_path: str) -> SecretStr:
    """
    Retrieve the secret key (password) from the Excel DB.

    Args:
        db_path (str): The path to the Excel db.

    Returns:
        SecretStr: The secret key.
    """
    try:
        admin_df = pd.read_excel(db_path, sheet_name="Admin")

        secret_key = admin_df.iloc[0]["password"]

        return SecretStr(secret_key)
    except IndexError:
        raise ValueError(f"Unable to get secret key from database.")
    except Exception as e:
        raise RuntimeError(f"Error reading secret key: {e}")


def create_schema_with_exclusions(
    schema_name: str,
    base_schema: Type[BaseModel],
    excluding_fields: List[str],
    optional: bool = False,
) -> Type[BaseModel]:
    """
    Dynamically create a new schema excluding specific fields from the base schema.
    """

    schema_fields = {}

    for k, v in base_schema.model_fields.items():
        if k not in excluding_fields:
            field_type = Optional[v.annotation] if optional else v.annotation
            field_default = None if optional else v.default

            schema_fields[k] = (field_type, field_default)

    new_schema = create_model(schema_name, **schema_fields)

    return new_schema


def read_from_db(db_path: str, sheet_name: str) -> Any:

    return pd.read_excel(db_path, sheet_name=sheet_name).dropna(how="all").fillna("")


def add_to_db(item: Any, db_path: str, sheet_name: str):
    workbook = load_workbook(db_path)
    sheet = workbook[sheet_name]

    first_empty_row = next(  # Get the first empty row by checking if cells are empty
        (
            row_i
            for row_i, row in enumerate(
                sheet.iter_rows(min_row=2, values_only=True), start=2
            )
            if all(cell is None for cell in row)
        ),
        sheet.max_row + 1,
    )

    # Handle both Pydantic models and dictionaries (needed to solve the save project data without tags to db)
    if hasattr(item, "__annotations__"):  # Check if it's a Pydantic model
        item_data = [
            (
                str(getattr(item, field))
                if isinstance(getattr(item, field), UUID)
                else getattr(item, field)
            )
            for field in item.__annotations__
        ]
    elif isinstance(item, dict):  # Handle dictionary input
        item_data = [
            str(value) if isinstance(value, UUID) else value for value in item.values()
        ]
    else:
        raise TypeError("Unsupported item type for add_to_db")

    for col_i, value in enumerate(item_data, start=1):
        value = (
            ", ".join([str(uuid) for uuid in value])
            if isinstance(value, list)
            else value
        )
        sheet.cell(row=first_empty_row, column=col_i, value=value)

    workbook.save(db_path)


def patch_db_entry(id_: str, patch_data: Any, db_path: str, sheet_name: str):
    data = read_from_db(db_path, sheet_name)

    item_to_patch = data[data["id"] == id_]

    if item_to_patch.empty:
        raise ValueError(f"Item with ID '{id_}' not found in the database.")

    for field, value in patch_data.dict(exclude_unset=True).items():
        if field in data.columns:
            value = (
                ", ".join([str(uuid) for uuid in value])
                if isinstance(value, list)
                else value
            )
            data.loc[data["id"] == id_, field] = value

    with pd.ExcelWriter(
        db_path, engine="openpyxl", mode="a", if_sheet_exists="replace"
    ) as writer:
        data.to_excel(writer, sheet_name=sheet_name, index=False)


# TODO - This delete leaves empty rows
def delete_from_db(id_: str, db_path: str, sheet_name: str):
    workbook = load_workbook(db_path)
    sheet = workbook[sheet_name]

    for row in sheet.iter_rows(min_row=2, values_only=False):
        if row[0].value == id_:
            for cell in row:
                cell.value = None
            break

    workbook.save(db_path)


# TODO - Merge the 2 delete from db methods
def delete_relation_from_db(
    project_id: str, tag_id: str, db_path: str, sheet_name: str
):
    data = read_from_db(db_path, sheet_name)

    data_to_delete = data[
        (data["project_id"] == project_id) & (data["tag_id"] == tag_id)
    ]

    if data_to_delete.empty:
        raise ValueError(
            f"Relationship with project_id '{project_id}' and tag_id '{tag_id}' not found."
        )

    data = data.drop(data_to_delete.index)

    with pd.ExcelWriter(
        db_path, engine="openpyxl", mode="a", if_sheet_exists="replace"
    ) as writer:
        data.to_excel(writer, sheet_name=sheet_name, index=False)
