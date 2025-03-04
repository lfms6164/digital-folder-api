from typing import Any, Union
from uuid import UUID

import pandas as pd
from fastapi import HTTPException
from openpyxl.reader.excel import load_workbook
from pydantic import BaseModel

from digital_folder.core.config import project_settings
from digital_folder.packages.Project.schemas import ProjectPatch
from digital_folder.packages.Tag.schemas import TagPatch


def read_from_db(db_path: str, sheet_name: str) -> Any:

    return pd.read_excel(db_path, sheet_name=sheet_name).dropna(how="all").fillna("")


def add_to_db(item: Any):
    sheet_map = {
        "TagOut": "Tags",
        "dict": "Projects",
        "RelationBase": "ProjectTagRelation",
    }

    sheet_name = sheet_map[type(item).__name__]

    wb = load_workbook(project_settings.EXCEL_DB_PATH)
    sheet = wb[sheet_name]

    first_empty_row = next(  # Get the first empty row by checking if cells are empty
        (
            row_i
            for row_i, row in enumerate(
                sheet.iter_rows(min_row=2, values_only=True), start=2
            )
            if all(cell is None for cell in row)
        ),
        sheet.max_row + 1,
    )

    item_dict = item.model_dump() if isinstance(item, BaseModel) else item

    item_data = [
        str(value) if isinstance(value, UUID) else value for value in item_dict.values()
    ]

    for col_i, value in enumerate(item_data, start=1):
        sheet.cell(row=first_empty_row, column=col_i, value=value)

    wb.save(project_settings.EXCEL_DB_PATH)
    wb.close()


def patch_db(id_: str, patch_data: Union[ProjectPatch, TagPatch]):
    sheet_map = {
        "TagPatch": "Tags",
        "ProjectPatch": "Projects",
    }

    sheet_name = sheet_map[type(patch_data).__name__]

    wb = load_workbook(project_settings.EXCEL_DB_PATH)
    sheet = wb[sheet_name]

    sheet_header_map = {header.value: i + 1 for i, header in enumerate(sheet[1])}

    row_to_patch = None
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, values_only=False):
        if str(row[0].value) == id_:
            row_to_patch = row
            break

    if not row_to_patch:
        raise HTTPException(
            status_code=400, detail=f"Patch failed - Item {id_} not found."
        )

    for field, value in patch_data.dict(exclude_unset=True).items():
        if field in sheet_header_map:
            row_to_patch[sheet_header_map[field] - 1].value = value

    wb.save(project_settings.EXCEL_DB_PATH)
    wb.close()


def delete_from_db(**ids):
    sheet_map = {
        ("tag_id",): "Tags",
        ("project_id",): "Projects",
        ("project_id", "tag_id"): "ProjectTagRelation",
    }

    ids_keys: Any = tuple(sorted(ids.keys()))

    if ids_keys not in sheet_map:
        raise HTTPException(
            status_code=400, detail=f"Delete failed - Invalid IDs provided."
        )

    sheet_name = sheet_map[ids_keys]

    wb = load_workbook(project_settings.EXCEL_DB_PATH)
    sheet = wb[sheet_name]

    row_to_delete = None
    for row in range(2, sheet.max_row + 1):
        if all(
            str(sheet.cell(row=row, column=i + 1).value) == ids[key]
            for i, key in enumerate(ids_keys)
        ):
            row_to_delete = row
            break

    if row_to_delete:
        sheet.delete_rows(row_to_delete)

    wb.save(project_settings.EXCEL_DB_PATH)
    wb.close()
